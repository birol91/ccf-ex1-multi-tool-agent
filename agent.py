"""
Exercise 1 — Multi-Tool Support Agent with Escalation Logic (CCA-F)

A customer-support agent built as a manual agentic loop. Built up across five steps:

  Step 1  Define 4 tools with clear, non-overlapping descriptions.
  Step 2  The stop_reason-driven agentic loop.
  Step 3  execute_tool() + structured errors (Excel mock database).
  Step 4  tool_hook() — confirm/escalate guardrail for refunds.
  Step 5  Multi-concern test scenarios (see __main__).

Model: claude-sonnet-4-6 · plain Anthropic SDK · no framework (manual loop needed for the hook).
"""

import json
import os

from anthropic import Anthropic
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load ANTHROPIC_API_KEY from the project-root .env (two levels up from this file).
load_dotenv(os.path.join(os.path.dirname(__file__), "..", "..", ".env"))

client = Anthropic()           # reads ANTHROPIC_API_KEY from the environment
MODEL = "claude-sonnet-4-6"

# System prompt — tells the model HOW to react to error categories (Step 3).
SYSTEM = (
    "You are a customer-support agent.\n"
    "- If a tool returns a 'transient' error with isRetryable=true, retry the same call once.\n"
    "- If it returns a 'validation' or 'permission' error, do NOT retry; explain the issue "
    "to the user in plain language.\n"
    "- Only perform irreversible actions (like refunds) when appropriate and explicitly "
    "requested.\n"
)

# ---------------------------------------------------------------------------
# Step 1 — Tool definitions
#
# These are just *descriptions* (a menu). Claude picks from this list; it never runs a tool
# itself — our code does (Step 3). Tool selection is driven entirely by `description`, so the
# two read-vs-write tools below MUST be clearly separated with boundary sentences, otherwise
# the model could route a harmless status query to an irreversible refund.
# ---------------------------------------------------------------------------

TOOLS = [
    {
        "name": "get_customer",
        "description": (
            "Fetch a customer's account info (name, membership tier, status). "
            "Read-only; changes nothing. "
            "Call this FIRST to verify identity or gather context. "
            "Do NOT use this for order or refund information — use lookup_order for orders."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "customer_id": {
                    "type": "string",
                    "description": "Customer ID, e.g. CUST-123",
                }
            },
            "required": ["customer_id"],
        },
    },
    {
        "name": "lookup_order",
        "description": (
            "Fetch the details of a single order (items, amount, status, date). "
            "Read-only: it ONLY shows information and does NOT move any money. "
            "Use this whenever the customer asks about an order's status, amount, or contents. "
            "This is NOT a refund — it never changes the order. "
            "To actually issue a refund, use process_refund instead."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "order_id": {
                    "type": "string",
                    "description": "Order ID, e.g. ORD-987",
                }
            },
            "required": ["order_id"],
        },
    },
    {
        "name": "process_refund",
        "description": (
            "Issue a MONEY REFUND against an order. This has SIDE EFFECTS and is IRREVERSIBLE. "
            "Call this ONLY when the customer explicitly asks for a refund AND the order is "
            "eligible. "
            "If the customer only wants to see order info (status, amount, contents), use "
            "lookup_order — NOT this tool. "
            "Never call this to merely display information."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "order_id": {"type": "string", "description": "Order ID to refund, e.g. ORD-987"},
                "amount": {"type": "number", "description": "Refund amount in the order's currency"},
            },
            "required": ["order_id", "amount"],
        },
    },
    {
        "name": "escalate_to_human",
        "description": (
            "Escalate the issue to a human agent. Use for cases that cannot be resolved "
            "automatically, that require policy approval, or that exceed an allowed threshold."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "reason": {"type": "string", "description": "Why this is being escalated"},
                "customer_id": {"type": "string", "description": "Customer ID, e.g. CUST-123"},
            },
            "required": ["reason", "customer_id"],
        },
    },
]


# ---------------------------------------------------------------------------
# Step 3 — Mock "database" (Excel) + structured errors + execute_tool
#
# Data lives in database.xlsx (built by create_database.py). In the real world this would be
# a database or an MCP server. Every tool returns the same envelope: a dict with
# "is_error" (bool) and "content" (a JSON string), so the loop in Step 2 can handle it uniformly.
# ---------------------------------------------------------------------------

DB_PATH = os.path.join(os.path.dirname(__file__), "database.xlsx")


def _read_sheet(sheet_name):
    """Read one Excel sheet into a list of dicts keyed by the header row."""
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(
            "database.xlsx not found — run `python create_database.py` first."
        )
    wb = load_workbook(DB_PATH, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return [dict(zip(header, row)) for row in rows[1:]]


def _find(sheet_name, key_field, key_value):
    for record in _read_sheet(sheet_name):
        if record[key_field] == key_value:
            return record
    return None


def make_error(category, retryable, message):
    """Structured error envelope (errorCategory / isRetryable / human-readable description)."""
    return {
        "is_error": True,
        "content": json.dumps({
            "errorCategory": category,      # "transient" | "validation" | "permission"
            "isRetryable": retryable,
            "description": message,
        }),
    }


def _ok(payload):
    return {"is_error": False, "content": json.dumps(payload)}


# Tracks which orders have already hit (and cleared) their one-time transient error.
# This makes ORD-FLAKY fail on the FIRST refund attempt and succeed on the retry — a
# deterministic way to exercise the retry path without real randomness.
_transient_cleared = set()


def execute_tool(name, args):
    if name == "get_customer":
        cust = _find("customers", "customer_id", args["customer_id"])
        if cust is None:
            return make_error("validation", False, f"Unknown customer: {args['customer_id']}")
        return _ok(cust)

    if name == "lookup_order":
        order = _find("orders", "order_id", args["order_id"])
        if order is None:
            return make_error("validation", False, f"Unknown order: {args['order_id']}")
        # Attach the line items and derive the total amount (sum of qty * unit_price)
        # rather than trusting a hand-entered amount column — this mirrors real systems.
        items = [
            {"product": r["product"], "qty": r["qty"], "unit_price": r["unit_price"]}
            for r in _read_sheet("order_items")
            if r["order_id"] == args["order_id"]
        ]
        order["items"] = items
        order["amount"] = sum(i["qty"] * i["unit_price"] for i in items)
        return _ok(order)

    if name == "process_refund":
        order = _find("orders", "order_id", args["order_id"])
        if order is None:
            return make_error("validation", False, f"Unknown order: {args['order_id']}")
        # Deterministic transient error: ORD-FLAKY fails once, then succeeds on retry.
        if args["order_id"] == "ORD-FLAKY" and args["order_id"] not in _transient_cleared:
            _transient_cleared.add(args["order_id"])
            return make_error("transient", True, "Refund service temporarily unavailable.")
        return _ok({"refunded": True, "order_id": args["order_id"], "amount": args["amount"]})

    if name == "escalate_to_human":
        return _ok({"escalated": True, "ticket": "ESC-001", "reason": args.get("reason", "")})

    return make_error("validation", False, f"Unknown tool: {name}")


# ---------------------------------------------------------------------------
# Step 4 — Programmatic hook (the confirm/escalate guardrail)
#
# Runs BEFORE a tool executes. A business rule enforced in CODE, not in the prompt — the
# model can be talked out of an instruction, but it cannot bypass code. The hook keys on the
# TOOL NAME alone (100% deterministic): no process_refund ever runs on the model's word.
#   - amount >= 500 : escalate to a human (large refunds need oversight).
#   - amount <  500 : ask the USER to confirm; only "yes" lets the refund proceed.
# Returning None means "no rule applies" → the loop runs execute_tool normally.
# ---------------------------------------------------------------------------

REFUND_THRESHOLD = 500.0


def tool_hook(name, args, confirm_callback=None):
    # `confirm_callback`: how the hook asks the user to confirm a small refund.
    #   - None  (default) → terminal mode: block on input() exactly as before.
    #   - callable        → web mode: delegate the yes/no decision to the caller.
    # The decision logic below is identical in both modes; only the *channel* changes,
    # so the terminal behaviour is untouched.
    if name != "process_refund":
        return None  # the gate only guards refunds; everything else flows normally

    amount = args.get("amount", 0)

    # Layer 1 — high-value refunds bypass automation entirely.
    if amount >= REFUND_THRESHOLD:
        return _ok({
            "blocked": True,
            "action": "escalate_to_human",
            "reason": f"Refund of {amount} >= {REFUND_THRESHOLD} requires human approval.",
        })

    # Layer 2 — smaller refunds need explicit user confirmation. Even if the model
    # mis-routed a status query into a refund, the user is asked before any money moves.
    if confirm_callback is None:
        answer = input(
            f"Confirm refund of {amount} for {args.get('order_id')}? (yes/no): "
        ).strip().lower()
        confirmed = answer == "yes"
    else:
        confirmed = confirm_callback(args)  # web mode: may raise ConfirmationRequired

    if not confirmed:
        return _ok({
            "blocked": True,
            "action": "cancelled",
            "reason": "User did not confirm the refund.",
        })

    return None  # confirmed → let execute_tool process the refund


# ---------------------------------------------------------------------------
# Step 2 — The agentic loop (the heart of the system)
#
# Drive the loop ONLY by response.stop_reason — never by parsing the text or by inspecting
# whether content has tool_use blocks. That is the official, reliable signal.
# ---------------------------------------------------------------------------

MAX_TURNS = 10   # safety net against an infinite loop — NOT the primary stop. The primary
                 # stop is always stop_reason == "end_turn".

def run_agent(user_message):
    """Run the agent on one user message and return the final text answer."""
    messages = [{"role": "user", "content": user_message}]

    for _ in range(MAX_TURNS):
        response = client.messages.create(
            model=MODEL,
            max_tokens=2048,
            system=SYSTEM,
            tools=TOOLS,
            messages=messages,
        )
        print(f"stop_reason: {response.stop_reason}")  # observe why the loop continues/stops

        # Append the assistant turn (including tool_use blocks) to history BEFORE handling
        # results. If we appended only the text, the tool_use ids would be lost and the
        # tool_result blocks could not match them (API 400).
        messages.append({"role": "assistant", "content": response.content})

        # Primary stop: the model has finished its answer.
        if response.stop_reason == "end_turn":
            return next((b.text for b in response.content if b.type == "text"), "")

        # The model wants to call one or more tools.
        if response.stop_reason == "tool_use":
            tool_results = []
            for block in response.content:
                if block.type == "tool_use":
                    # Step 4: a hook may intercept the call before it runs.
                    hooked = tool_hook(block.name, block.input)
                    result = hooked if hooked is not None else execute_tool(block.name, block.input)

                    tool_result = {
                        "type": "tool_result",
                        "tool_use_id": block.id,        # MUST match the tool_use id
                        "content": result["content"],
                    }
                    if result["is_error"]:
                        tool_result["is_error"] = True
                    tool_results.append(tool_result)

            # Feed ALL tool results back in ONE user message (supports parallel tool calls).
            messages.append({"role": "user", "content": tool_results})
            continue

        # Unexpected stop_reason (e.g. "max_tokens", "refusal"): don't loop blindly.
        print(f"[warning] unexpected stop_reason: {response.stop_reason}")
        return next((b.text for b in response.content if b.type == "text"), "")

    return "[stopped] reached MAX_TURNS safety net."


# ---------------------------------------------------------------------------
# Web entry point — same loop, no input()
#
# run_agent() above is the terminal version (blocking input() for confirmations) and is left
# untouched. For the browser UI we need two things it cannot give us:
#   1. multi-turn memory: the loop must accept and return the full `messages` history so the
#      conversation continues across HTTP requests;
#   2. in-chat confirmation: a small refund must NOT call input(); instead the loop pauses and
#      hands control back to the web layer, which asks the user in the chat and resumes later.
#
# We achieve (2) with a ConfirmationRequired exception raised from the hook's confirm_callback.
# It unwinds the loop cleanly, leaving `messages` at a safe point (the assistant turn that
# requested the refund has NOT been appended yet), so the saved history is always API-valid.
# On the next request the web layer re-runs run_agent_web with a pre-supplied decision.
# ---------------------------------------------------------------------------


class ConfirmationRequired(Exception):
    """Raised mid-loop when a small refund needs the user's yes/no in the chat."""
    def __init__(self, refund_args):
        super().__init__("refund confirmation required")
        self.refund_args = refund_args


def run_agent_web(messages, refund_decision=None):
    """
    Drive the agentic loop over an existing `messages` history (web/multi-turn).

    `refund_decision`:
        None  → if a small refund is hit, raise ConfirmationRequired (pause for the chat).
        True  → a refund the user already approved; let it proceed.
        False → a refund the user already declined; cancel it.
    The pre-supplied decision is consumed for the FIRST refund encountered, which is the one
    that was pending. Any further refund in the same run falls back to asking again.

    Returns (reply_text, messages) on completion. Raises ConfirmationRequired to pause.
    """
    decision_box = {"value": refund_decision}

    def confirm_callback(_args):
        if decision_box["value"] is not None:
            answer = decision_box["value"]
            decision_box["value"] = None  # one-shot: only the pending refund uses it
            return answer
        raise ConfirmationRequired(_args)

    for _ in range(MAX_TURNS):
        response = client.messages.create(
            model=MODEL,
            max_tokens=2048,
            system=SYSTEM,
            tools=TOOLS,
            messages=messages,
        )

        if response.stop_reason == "end_turn":
            messages.append({"role": "assistant", "content": response.content})
            reply = next((b.text for b in response.content if b.type == "text"), "")
            return reply, messages

        if response.stop_reason == "tool_use":
            # Run the hook for every refund block FIRST. If one needs confirmation, raise
            # BEFORE appending this assistant turn — so `messages` stays API-valid for resume.
            tool_results = []
            for block in response.content:
                if block.type == "tool_use":
                    hooked = tool_hook(block.name, block.input, confirm_callback)
                    result = hooked if hooked is not None else execute_tool(block.name, block.input)
                    tool_result = {
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result["content"],
                    }
                    if result["is_error"]:
                        tool_result["is_error"] = True
                    tool_results.append(tool_result)

            messages.append({"role": "assistant", "content": response.content})
            messages.append({"role": "user", "content": tool_results})
            continue

        messages.append({"role": "assistant", "content": response.content})
        reply = next((b.text for b in response.content if b.type == "text"), "")
        return reply, messages

    return "[stopped] reached MAX_TURNS safety net.", messages


# ---------------------------------------------------------------------------
# Step 5 — Multi-concern test
#
# A single message carries TWO independent requests. The agent must decompose them, pick the
# right tool for each, and synthesise ONE combined reply. This exercises tool selection
# (Step 1), the loop (Step 2), structured errors (Step 3) and the hook (Step 4) all at once.
# Requires ANTHROPIC_API_KEY + the `anthropic` package + `python create_database.py`.
# ---------------------------------------------------------------------------

def _run_scenario(title, message):
    print("\n" + "=" * 70)
    print(title)
    print("=" * 70)
    print(f"USER: {message}\n")
    print(f"AGENT: {run_agent(message)}")


if __name__ == "__main__":
    # Scenario A — order status + a HIGH-value refund (>= 500).
    # Expected: get_customer → lookup_order (status + items) → process_refund(900) hits the
    # hook → escalated to a human. One combined reply covers both the status and the escalation.
    _run_scenario(
        "Scenario A — status + high-value refund (escalation path)",
        "Hi, I'm CUST-123. Two things: (1) what is the status of order ORD-987? "
        "(2) I also want a 900 refund on order ORD-654.",
    )

    # Scenario B — order status + a SMALL refund (< 500).
    # Expected: lookup_order → process_refund(300) hits the hook → asks YOU to confirm in the
    # terminal (yes/no). Try answering "no" to see the refund cancelled while the status is
    # still reported. This is the confirmation-gate path.
    _run_scenario(
        "Scenario B — status + small refund (confirmation path; will prompt yes/no)",
        "Hi, I'm CUST-123. Please tell me the status of order ORD-987, "
        "and also process a 300 refund on order ORD-FLAKY.",
    )
