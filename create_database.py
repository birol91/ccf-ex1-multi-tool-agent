"""
One-off helper: build the mock `database.xlsx` (customers + orders + order_items sheets).

The order total is NOT stored; lookup_order derives it from order_items (sum of
qty * unit_price), mirroring how real systems compute an order's amount.

Run once: `python create_database.py`. The agent reads this file at runtime.
In the real world this would be a database or an MCP server; here it's a mock Excel file.
"""

from openpyxl import Workbook

wb = Workbook()

# --- Sheet 1: customers ---
ws_c = wb.active
ws_c.title = "customers"
ws_c.append(["customer_id", "name", "tier", "status"])
ws_c.append(["CUST-123", "Ayse Yilmaz", "gold", "active"])
ws_c.append(["CUST-456", "Mehmet Demir", "silver", "active"])
ws_c.append(["CUST-789", "Zeynep Kaya", "bronze", "suspended"])

# --- Sheet 2: orders (header only; the total `amount` is derived from order_items) ---
ws_o = wb.create_sheet("orders")
ws_o.append(["order_id", "customer_id", "status"])
ws_o.append(["ORD-987", "CUST-123", "delivered"])     # items total 250  → < 500, confirm path
ws_o.append(["ORD-654", "CUST-123", "delivered"])     # items total 900  → >= 500, escalate path
ws_o.append(["ORD-321", "CUST-456", "shipped"])       # items total 120
ws_o.append(["ORD-FLAKY", "CUST-123", "delivered"])   # items total 300, transient error once

# --- Sheet 3: order_items (one row per product line; amount = sum(qty * unit_price)) ---
ws_i = wb.create_sheet("order_items")
ws_i.append(["order_id", "product", "qty", "unit_price"])
ws_i.append(["ORD-987", "Wireless Mouse", 2, 50.0])      # 100
ws_i.append(["ORD-987", "Mechanical Keyboard", 1, 150.0])  # 150  → ORD-987 total = 250
ws_i.append(["ORD-654", "27\" Monitor", 2, 450.0])         # 900  → ORD-654 total = 900
ws_i.append(["ORD-321", "USB-C Cable", 3, 40.0])           # 120  → ORD-321 total = 120
ws_i.append(["ORD-FLAKY", "Webcam", 1, 300.0])             # 300  → ORD-FLAKY total = 300

wb.save("database.xlsx")
print("database.xlsx created (customers: 3, orders: 4, order_items: 5).")
